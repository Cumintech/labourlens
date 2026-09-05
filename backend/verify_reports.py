"""Verifies Day 4's report generation + email delivery for real: downloads
a real Excel/PDF file and parses it back (not just checking the endpoint
returns 200), and spins up a REAL local SMTP server (aiosmtpd) to receive
an actual SMTP session -- not a mocked smtplib.SMTP. Also confirms an
unreachable SMTP server surfaces as a real error rather than silently
"succeeding".

Requires aiosmtpd: pip install -r requirements-dev.txt

    DATABASE_URL=sqlite:///./scratch.db JWT_SECRET=x ENCRYPTION_KEY=<fernet key> python verify_reports.py
"""

import email
import io
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from datetime import date

from aiosmtpd.controller import Controller
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from database import Base, engine
import models
from main import app

Base.metadata.create_all(bind=engine)
client = TestClient(app)
TODAY = date(2026, 8, 18)
START = date(2026, 8, 1)


class CapturingHandler:
    def __init__(self):
        self.messages = []

    async def handle_DATA(self, server, session, envelope):
        self.messages.append(
            {"mail_from": envelope.mail_from, "rcpt_tos": envelope.rcpt_tos, "content": envelope.content}
        )
        return "250 Message accepted for delivery"


handler = CapturingHandler()
# Port 1026, not the dev-relay's usual 1025 -- this test's own SMTP
# server must never fight a real dev_smtp_relay.py that's already
# running for manual/real-device testing (a genuine port-bind conflict
# hit while running the full verify suite alongside a live dev session).
controller = Controller(handler, hostname="127.0.0.1", port=1026)
controller.start()
# Override every SMTP_* var explicitly rather than relying on whatever
# .env currently has configured (a real provider like Gmail, with TLS
# on) -- this test always talks to the local plain-SMTP controller above.
prior_smtp_env = {k: os.environ.get(k) for k in ("SMTP_HOST", "SMTP_PORT", "SMTP_USE_TLS", "SMTP_USER", "SMTP_PASSWORD")}
os.environ["SMTP_HOST"] = "127.0.0.1"
os.environ["SMTP_PORT"] = "1026"
os.environ["SMTP_USE_TLS"] = "false"
os.environ.pop("SMTP_USER", None)
os.environ.pop("SMTP_PASSWORD", None)

try:
    signup = client.post(
        "/owners/signup",
        json={"name": "Report Owner", "mobile": "9000000075", "password": "pass123", "factory_name": "Report Factory"},
    )
    assert signup.status_code == 201, signup.text
    token = signup.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    w = client.post("/workers", headers=headers, json={"name": "Report Worker", "aadhaar_number": "999988887777"}).json()
    mark = client.post(
        "/attendance", headers=headers, json={"worker_id": w["id"], "date": str(TODAY), "slot": "AM", "status": "present"}
    )
    assert mark.status_code == 200, mark.text

    # --- Direct download: Excel -- real file, parsed back with openpyxl ---
    resp = client.get(
        "/reports/attendance", headers=headers, params={"start_date": str(START), "end_date": str(TODAY), "format": "excel"}
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats"), resp.headers
    wb = load_workbook(io.BytesIO(resp.content))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0] == ("Worker", "Aadhaar (last 4)", "Date", "Slot", "Status"), rows[0]
    assert any(r[0] == "Report Worker" and r[3] == "AM" and r[4] == "present" for r in rows[1:]), rows
    print("Excel report download: real file, parseable, correct data: PASSED")

    # --- Direct download: PDF -- real file, magic bytes + size checked ---
    resp = client.get(
        "/reports/attendance", headers=headers, params={"start_date": str(START), "end_date": str(TODAY), "format": "pdf"}
    )
    assert resp.status_code == 200, resp.text
    assert resp.content[:4] == b"%PDF", "response is not a real PDF"
    assert len(resp.content) > 500, "PDF suspiciously small"
    print("PDF report download: real PDF magic bytes, non-trivial size: PASSED")

    # --- Email delivery: real SMTP session against the real local server above ---
    resp = client.post(
        "/reports/attendance/email",
        headers=headers,
        json={"start_date": str(START), "end_date": str(TODAY), "recipient_email": "owner@example.com", "format": "excel"},
    )
    assert resp.status_code == 202, resp.text
    assert len(handler.messages) == 1, f"expected exactly one email received by the real SMTP server: {handler.messages}"
    msg = handler.messages[0]
    assert msg["rcpt_tos"] == ["owner@example.com"], msg["rcpt_tos"]

    parsed = email.message_from_bytes(msg["content"])
    attachments = [p for p in parsed.walk() if p.get_filename()]
    assert len(attachments) == 1, f"expected exactly one attachment: {attachments}"
    attachment_bytes = attachments[0].get_payload(decode=True)
    wb2 = load_workbook(io.BytesIO(attachment_bytes))
    rows2 = list(wb2.active.iter_rows(values_only=True))
    assert any(r[0] == "Report Worker" for r in rows2[1:]), "attachment content doesn't match the report"
    print("Email delivered over a real SMTP session, with a real parseable attachment: PASSED")

    # --- Genuine failure path: SMTP server unreachable ---
    os.environ["SMTP_PORT"] = "1099"  # nothing listening here
    client_no_raise = TestClient(app, raise_server_exceptions=False)
    client_no_raise.headers.update(headers)
    resp = client_no_raise.post(
        "/reports/attendance/email",
        json={"start_date": str(START), "end_date": str(TODAY), "recipient_email": "owner@example.com", "format": "pdf"},
    )
    assert resp.status_code == 500, f"unreachable SMTP server should surface as a real error, not silently succeed: {resp.status_code}"
    os.environ["SMTP_PORT"] = "1026"
    print("unreachable SMTP server surfaces as a real error, not silently swallowed: PASSED")

finally:
    for key, value in prior_smtp_env.items():
        if value is None:
            os.environ.pop(key, None)
        else:
            os.environ[key] = value
    controller.stop()

print("\nALL ASSERTIONS PASSED")
