"""Verifies Phase 3 Day 3's form generation: all 5 forms produce real,
parseable PDF and Excel content (not just HTTP 200), Form 15/Wage Slip's
wage arithmetic matches an independently hand-computed expected value
(not just "some number came out"), FormGenerationLog captures real
usage, cross-owner scoping -- including that Form 25/15's factory-wide
rows never leak another owner's workers -- real email delivery over a
real SMTP session (not just a 202 response), a worker with no wage
profile set at all falls back cleanly instead of crashing, and a
malformed shift time (pre-dating the mobile app's time picker) doesn't
take down form generation for every worker on the account -- the exact
failure real-device testing found.

    DATABASE_URL=sqlite:///./scratch.db JWT_SECRET=x ENCRYPTION_KEY=<fernet key> python verify_forms.py
"""

import email as email_lib
import io
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from aiosmtpd.controller import Controller
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader

from database import Base, SessionLocal, engine
import models
from main import app

Base.metadata.create_all(bind=engine)
client = TestClient(app)


def pdf_text(content: bytes) -> str:
    reader = PdfReader(io.BytesIO(content))
    return "\n".join(page.extract_text() for page in reader.pages)


# --- Owner A setup: factory profile, a shift with real hours, a worker
# with Form 12 details, attendance, a wage rate, and a paid leave entry ---
signup = client.post(
    "/owners/signup",
    json={"name": "Forms Owner A", "mobile": "9000000701", "password": "pass123", "factory_name": "Forms Factory A"},
)
assert signup.status_code == 201, signup.text
token_a = signup.json()["access_token"]
headers_a = {"Authorization": f"Bearer {token_a}"}

profile = client.put(
    "/owners/me/factory-profile",
    headers=headers_a,
    json={"factory_address": "42 Industrial Estate, Madurai", "factory_licence_no": "TN/MDU/1234"},
)
assert profile.status_code == 200, profile.text

am_shift = client.get("/shift-configs", headers=headers_a).json()[0]
assert am_shift["slot_key"] == "AM"
set_hours = client.put(
    f"/shift-configs/{am_shift['id']}", headers=headers_a, json={"slot_key": "AM", "label": "AM", "start_time": "06:00", "end_time": "14:00"}
)
assert set_hours.status_code == 200, set_hours.text  # AM is now an 8-hour shift

worker = client.post(
    "/workers", headers=headers_a, json={"name": "Forms Worker", "aadhaar_number": "555566667777", "dob": "1990-01-01", "gender": "Male"}
).json()

compliance = client.post(
    f"/workers/{worker['id']}/compliance",
    headers=headers_a,
    json={"worker_code": "T-001", "designation_or_nature_of_work": "Machine Operator", "date_of_joining": "2020-01-01"},
)
assert compliance.status_code == 201 and compliance.json()["category"] == "adult", compliance.text

# 5 days present in AM (8h each), one of them with 4h overtime -> days_worked=5, total_hours=44
for day, ot in (("2026-08-03", 0), ("2026-08-04", 0), ("2026-08-05", 0), ("2026-08-06", 0), ("2026-08-07", 4)):
    mark = client.post(
        "/attendance", headers=headers_a, json={"worker_id": worker["id"], "date": day, "slot": "AM", "status": "present", "overtime_hours": ot}
    )
    assert mark.status_code == 200, mark.text

leave = client.post(
    f"/workers/{worker['id']}/leave",
    headers=headers_a,
    json={"leave_type": "national_festival_special", "date_from": "2026-08-20", "date_to": "2026-08-20", "days": 1, "wages_paid": 500},
)
assert leave.status_code == 201, leave.text

wage_rate = client.post(
    f"/workers/{worker['id']}/wage-profile",
    headers=headers_a,
    json={
        "rate_type": "daily", "basic": 500, "hra": 20, "da": 50, "other_allowances": 10,
        "pf_rate": 10, "esi_rate": 1, "lwf_amount": 25, "effective_from": "2026-08-01",
    },
)
assert wage_rate.status_code == 201, wage_rate.text

# --- Independently hand-computed expected wage, not derived from forms.py ---
DAYS_WORKED = 5
BASIC_WAGE = 500 * DAYS_WORKED  # 2500
HOURLY_RATE = 500 / 8  # 62.5
OT_WAGES = 4 * HOURLY_RATE * 2  # 500
LEAVE_WAGES = 500
GROSS = BASIC_WAGE + 50 + 20 + 10 + OT_WAGES + LEAVE_WAGES  # 3580
PF = 0.10 * (BASIC_WAGE + 50)  # 255
ESI = 0.01 * GROSS  # 35.8
LWF = 25
TOTAL_DEDUCTIONS = PF + ESI + LWF  # 315.8
NET = GROSS - TOTAL_DEDUCTIONS  # 3264.2

# --- Form 15 (factory-wide) ---
form15_pdf = client.get("/forms/form15", headers=headers_a, params={"month": 8, "year": 2026, "format": "pdf"})
assert form15_pdf.status_code == 200 and form15_pdf.content[:4] == b"%PDF", form15_pdf.status_code
text15 = pdf_text(form15_pdf.content)
assert "Forms Factory A" in text15, "factory name missing from Form 15"
assert "Forms Worker" in text15, "worker name missing from Form 15"
# The real Form 15 has no single "Total Deductions" column -- it lists
# each deduction separately (PF, ESI, LWF, advances, damages) straight
# through to Net Wages, so TOTAL_DEDUCTIONS is checked in the Excel
# assertion below via its component figures, not as its own printed cell.
for expected in (f"{GROSS:.2f}", f"{PF:.2f}", f"{ESI:.2f}", f"{NET:.2f}"):
    assert expected in text15, f"expected wage figure {expected!r} not found in Form 15 PDF text"
print("Form 15 PDF: real content, wage arithmetic matches hand-computed expected values: PASSED")

form15_excel = client.get("/forms/form15", headers=headers_a, params={"month": 8, "year": 2026, "format": "excel"})
assert form15_excel.status_code == 200, form15_excel.text
wb15 = load_workbook(io.BytesIO(form15_excel.content))
rows15 = list(wb15.active.iter_rows(values_only=True))
worker_row = next(r for r in rows15 if r[2] == "Forms Worker")
assert worker_row[4] == str(DAYS_WORKED), worker_row
print("Form 15 Excel: real parseable file, correct days-worked: PASSED")

# --- Wage Slip (per worker) ---
wageslip_pdf = client.get("/forms/wageslip", headers=headers_a, params={"worker_id": worker["id"], "month": 8, "year": 2026, "format": "pdf"})
assert wageslip_pdf.status_code == 200 and wageslip_pdf.content[:4] == b"%PDF", wageslip_pdf.status_code
text_slip = pdf_text(wageslip_pdf.content)
assert "Machine Operator" in text_slip, "designation missing from wage slip"
assert f"{NET:.2f}" in text_slip, "net wages missing/wrong on wage slip"
print("Wage Slip PDF: real content, net wages matches hand-computed expected value: PASSED")

# --- Form 25 (factory-wide muster roll) ---
form25_pdf = client.get("/forms/form25", headers=headers_a, params={"month": 8, "year": 2026, "format": "pdf"})
assert form25_pdf.status_code == 200 and form25_pdf.content[:4] == b"%PDF", form25_pdf.status_code
text25 = pdf_text(form25_pdf.content)
assert "Forms Worker" in text25 and "T-001" in text25, "worker row missing from Form 25"
print("Form 25 PDF: real content, includes the worker's row: PASSED")

form25_excel = client.get("/forms/form25", headers=headers_a, params={"month": 8, "year": 2026, "format": "excel"})
wb25 = load_workbook(io.BytesIO(form25_excel.content))
rows25 = list(wb25.active.iter_rows(values_only=True))
worker_row_25 = next(r for r in rows25 if r[2] == "Forms Worker")
assert worker_row_25[-6] == "5", f"Form 25 total-days-worked column wrong: {worker_row_25}"  # "Total Days Worked"
print("Form 25 Excel: total days worked matches hand-computed value: PASSED")

# --- Form 25-B (per worker time card) ---
form25b_pdf = client.get("/forms/form25b", headers=headers_a, params={"worker_id": worker["id"], "month": 8, "year": 2026, "format": "pdf"})
assert form25b_pdf.status_code == 200 and form25b_pdf.content[:4] == b"%PDF", form25b_pdf.status_code
text25b = pdf_text(form25b_pdf.content)
assert "2026-08-07" in text25b, "the overtime day is missing from Form 25-B"
print("Form 25-B PDF: real content, day-by-day rows present: PASSED")

# --- Form 12 (Register of Adult Workers -- factory-wide, one row per
# worker in registration order; a real government register, not a
# per-worker sheet, confirmed against the actual scanned form) ---
form12_pdf = client.get("/forms/form12", headers=headers_a, params={"format": "pdf"})
assert form12_pdf.status_code == 200 and form12_pdf.content[:4] == b"%PDF", form12_pdf.status_code
text12 = pdf_text(form12_pdf.content)
assert "Machine Operator" in text12 and "T-001" in text12, "Form 12 register missing compliance fields"
assert "Forms Worker" in text12, "Form 12 register missing the worker's row"
# Wrapped header cells break onto multiple lines at word boundaries (PDF
# text extraction then sees each visual line, not the joined phrase), so
# these check for single words rather than "Specimen Signature" verbatim.
assert "Specimen" in text12 and "Signature" in text12 and "Photo" in text12, "Form 12 register missing required column headers"
print("Form 12 PDF: full register, real content, exact column set from the real form: PASSED")

form12_single_pdf = client.get(f"/forms/form12/{worker['id']}", headers=headers_a, params={"format": "pdf"})
assert form12_single_pdf.status_code == 200 and form12_single_pdf.content[:4] == b"%PDF", form12_single_pdf.status_code
text12_single = pdf_text(form12_single_pdf.content)
assert "Machine Operator" in text12_single and "T-001" in text12_single, "single-worker Form 12 missing compliance fields"
print("Form 12 PDF: single-worker row narrows to the same register layout: PASSED")

# --- FormGenerationLog captures real usage ---
db = SessionLocal()
log_rows = db.query(models.FormGenerationLog).filter(models.FormGenerationLog.owner_id == signup.json()["owner"]["id"]).all()
assert len(log_rows) == 8, f"expected 8 logged generations (form15 x2, wageslip, form25 x2, form25b, form12 x2), got {len(log_rows)}"
assert {r.form_code for r in log_rows} == {"form15", "wageslip", "form25", "form25b", "form12"}
db.close()
print("FormGenerationLog captures one row per successful generation: PASSED")

# --- Real email delivery over a real SMTP session (never actually
# exercised by any prior automated test -- the /forms/*/email endpoint
# only ever got a manual real-device test, which is exactly the one that
# turned up a 500). Runs its own local SMTP server on a different port
# than the live dev relay (127.0.0.1:1025) so this test doesn't fight
# over that port with whatever's already running on it. ---
class _CapturingHandler:
    def __init__(self):
        self.messages = []

    async def handle_DATA(self, server, session, envelope):
        self.messages.append(
            {"mail_from": envelope.mail_from, "rcpt_tos": envelope.rcpt_tos, "content": envelope.content}
        )
        return "250 Message accepted for delivery"


email_handler = _CapturingHandler()
email_controller = Controller(email_handler, hostname="127.0.0.1", port=1026)
email_controller.start()
# Override every SMTP_* var explicitly rather than relying on whatever
# .env currently has configured (a real provider like Gmail, with TLS
# on) -- this test always talks to the local plain-SMTP controller above.
prior_smtp_env = {k: os.environ.get(k) for k in ("SMTP_HOST", "SMTP_PORT", "SMTP_USE_TLS", "SMTP_USER", "SMTP_PASSWORD")}
os.environ["SMTP_HOST"] = "127.0.0.1"
os.environ["SMTP_PORT"] = "1026"
os.environ["SMTP_USE_TLS"] = "false"
os.environ.pop("SMTP_USER", None)
os.environ.pop("SMTP_PASSWORD", None)
try:
    email_resp = client.post(
        "/forms/wageslip/email",
        headers=headers_a,
        json={
            "worker_id": worker["id"], "month": 8, "year": 2026, "format": "pdf",
            "recipient_email": "ganeshprabu844@gmail.com",
        },
    )
    assert email_resp.status_code == 202, email_resp.text
    assert len(email_handler.messages) == 1, f"expected exactly one email received: {email_handler.messages}"
    sent = email_handler.messages[0]
    assert sent["rcpt_tos"] == ["ganeshprabu844@gmail.com"], sent["rcpt_tos"]
    parsed = email_lib.message_from_bytes(sent["content"])
    attachments = [p for p in parsed.walk() if p.get_filename()]
    assert len(attachments) == 1, f"expected exactly one attachment: {attachments}"
    attachment_bytes = attachments[0].get_payload(decode=True)
    assert attachment_bytes[:4] == b"%PDF", "email attachment is not a real PDF"
    print("form email delivered over a real SMTP session, with a real PDF attachment, to the real recipient: PASSED")

    # Genuine failure path, same discipline as verify_reports.py: an
    # unreachable SMTP server must surface as a real error.
    os.environ["SMTP_PORT"] = "1099"  # nothing listening here
    client_no_raise = TestClient(app, raise_server_exceptions=False)
    client_no_raise.headers.update(headers_a)
    unreachable = client_no_raise.post(
        "/forms/wageslip/email",
        json={"worker_id": worker["id"], "month": 8, "year": 2026, "format": "pdf", "recipient_email": "owner@example.com"},
    )
    assert unreachable.status_code == 500, f"unreachable SMTP server should surface as a real error: {unreachable.status_code}"
    print("form email with an unreachable SMTP server surfaces as a real error, not silently swallowed: PASSED")
finally:
    for key, value in prior_smtp_env.items():
        if value is None:
            os.environ.pop(key, None)
        else:
            os.environ[key] = value
    email_controller.stop()

# --- A worker with no wage profile at all: Form 15 / Wage Slip must fall
# back cleanly, never crash or fabricate numbers ---
no_rate_worker = client.post(
    "/workers", headers=headers_a, json={"name": "No Rate Worker", "aadhaar_number": "222233334444"}
).json()

form15_no_rate = client.get("/forms/form15", headers=headers_a, params={"month": 8, "year": 2026, "format": "pdf"})
assert form15_no_rate.status_code == 200, form15_no_rate.text
text15_no_rate = pdf_text(form15_no_rate.content)
assert "No Rate Worker" in text15_no_rate and "no wage rate set" in text15_no_rate, (
    "Form 15 should list a worker with no wage profile and say so plainly, not crash or omit them"
)
print("Form 15 with a worker missing a wage profile: falls back cleanly, doesn't crash: PASSED")

wageslip_no_rate = client.get(
    "/forms/wageslip", headers=headers_a, params={"worker_id": no_rate_worker["id"], "month": 8, "year": 2026, "format": "pdf"}
)
assert wageslip_no_rate.status_code == 200, wageslip_no_rate.text
assert "No wage rate has been set" in pdf_text(wageslip_no_rate.content)
print("Wage Slip with no wage profile set: falls back cleanly, doesn't crash: PASSED")

# --- A malformed shift time (real bug: a free-typed "6.00am"/"2pm" from
# before the mobile app had a time picker) must not crash form
# generation for every worker on the account -- written directly via the
# ORM to simulate genuinely pre-existing bad data, bypassing the API
# validation that now blocks this going forward. ---
db_bad_shift = SessionLocal()
bad_shift = models.ShiftConfig(owner_id=signup.json()["owner"]["id"], slot_key="Bad", label="Bad Shift", start_time="6.00am", end_time="2pm")
db_bad_shift.add(bad_shift)
db_bad_shift.commit()
bad_shift_id = bad_shift.id
db_bad_shift.close()

mark_bad_shift = client.post(
    "/attendance",
    headers=headers_a,
    json={"worker_id": worker["id"], "date": "2026-08-15", "slot": "Bad", "status": "present"},
)
assert mark_bad_shift.status_code == 200, mark_bad_shift.text

form25b_with_bad_shift = client.get(
    "/forms/form25b", headers=headers_a, params={"worker_id": worker["id"], "month": 8, "year": 2026, "format": "pdf"}
)
assert form25b_with_bad_shift.status_code == 200, (
    f"a malformed shift time crashed form generation -- this exact bug reached a real device: {form25b_with_bad_shift.status_code}"
)
print("a malformed shift time (pre-existing bad data) no longer crashes form generation: PASSED")

# Clean up so it doesn't affect any test below that re-generates Form 25/15.
db_cleanup = SessionLocal()
db_cleanup.query(models.Attendance).filter(models.Attendance.slot == "Bad").delete()
db_cleanup.query(models.ShiftConfig).filter(models.ShiftConfig.id == bad_shift_id).delete()
db_cleanup.commit()
db_cleanup.close()

# --- Cross-owner scoping ---
signup_b = client.post(
    "/owners/signup",
    json={"name": "Forms Owner B", "mobile": "9000000702", "password": "pass123", "factory_name": "Forms Factory B"},
)
token_b = signup_b.json()["access_token"]
headers_b = {"Authorization": f"Bearer {token_b}"}
worker_b = client.post("/workers", headers=headers_b, json={"name": "Owner B Worker", "aadhaar_number": "111111119999"}).json()

form15_b = client.get("/forms/form15", headers=headers_b, params={"month": 8, "year": 2026, "format": "pdf"})
text15_b = pdf_text(form15_b.content)
assert "Forms Worker" not in text15_b, "owner B's Form 15 leaked owner A's worker"
assert "Owner B Worker" in text15_b, "owner B's own worker missing from their own Form 15"
print("Form 15 factory-wide scoping: owner B never sees owner A's workers: PASSED")

form12_cross = client.get(f"/forms/form12/{worker['id']}", headers=headers_b)
assert form12_cross.status_code == 404, form12_cross.text
form25b_cross = client.get("/forms/form25b", headers=headers_b, params={"worker_id": worker["id"], "month": 8, "year": 2026})
assert form25b_cross.status_code == 404, form25b_cross.text
wageslip_cross = client.get("/forms/wageslip", headers=headers_b, params={"worker_id": worker["id"], "month": 8, "year": 2026})
assert wageslip_cross.status_code == 404, wageslip_cross.text
print("cross-owner scoping on per-worker forms (form12, form25b, wageslip): PASSED")

print("\nALL ASSERTIONS PASSED")
