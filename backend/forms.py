"""Phase 3 Day 3 -- statutory form generation (Form 25, 25-B, 12, 15,
Wage Slip). Sibling to reports.py, reusing its reportlab/openpyxl setup,
but reports.py has no header/footer helper to reuse -- it's a single
title Paragraph + one Table. _header_elements() below is new work, not
a reuse of something that already existed.

Form 25 and Form 15 are factory-wide registers (every worker as a row,
one document per month) -- confirmed from the real Tamil Nadu form
PDFs. Form 25-B, Form 12, and Wage Slip are per-worker. See
PHASE3_STATUTORY_FORMS_PLAN.md for the full field-by-field basis.

No cached rollups: every figure below is computed at request time from
Attendance / LeaveEntry / WageProfile / WagePayment, nothing new is
written to represent "the computed result" (ground rule carried over
from reports.py's 6-month report).
"""

import calendar
import io
from datetime import date, timedelta

from openpyxl import Workbook
from reportlab.lib import colors as pdf_colors
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

import models

# Neither of these is a confirmed current statutory figure -- both are
# assumptions flagged in PHASE3_STATUTORY_FORMS_PLAN.md's Day 3 section.
# STANDARD_DAILY_HOURS backs out an hourly rate from a daily/monthly wage
# for the OT calculation; OT_MULTIPLIER is the Factories Act Sec. 59
# standard basis. Confirm both against the real current rules before
# relying on generated wage numbers for actual payroll/filing.
STANDARD_DAILY_HOURS = 8
OT_MULTIPLIER = 2

PAID_LEAVE_TYPES = ("earned", "national_festival_special")


# --------------------------------------------------------------------------
# Shared data helpers -- used by more than one form, so the arithmetic
# (and any bug in it) lives in exactly one place.
# --------------------------------------------------------------------------


def _month_date_range(month: int, year: int) -> tuple[date, date]:
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def _shift_map(db: Session, owner_id: int) -> dict[str, models.ShiftConfig]:
    shifts = db.query(models.ShiftConfig).filter(models.ShiftConfig.owner_id == owner_id).all()
    return {s.slot_key: s for s in shifts}


def _shift_duration_hours(shift: models.ShiftConfig | None) -> float:
    """Form 25/25-B's daily hours are the shift's standard duration
    whenever marked present, not a real clocked time -- see Day 1's
    confirmed v1 approach in PHASE3_STATUTORY_FORMS_PLAN.md. Handles an
    overnight shift (e.g. 22:00-06:00) by wrapping past midnight."""
    if shift is None or not shift.start_time or not shift.end_time:
        return 0.0
    try:
        sh, sm = (int(p) for p in shift.start_time.split(":"))
        eh, em = (int(p) for p in shift.end_time.split(":"))
    except (ValueError, TypeError):
        # A malformed stored time (e.g. pre-dating the mobile app's move
        # to a real time picker) must never take down form generation for
        # every worker -- treat it as "duration unknown" instead of
        # crashing the whole request.
        return 0.0
    start_minutes = sh * 60 + sm
    end_minutes = eh * 60 + em
    if end_minutes <= start_minutes:
        end_minutes += 24 * 60
    return (end_minutes - start_minutes) / 60


def _paid_leave_dates(db: Session, worker_id: int, start_date: date, end_date: date) -> set[date]:
    entries = (
        db.query(models.LeaveEntry)
        .filter(
            models.LeaveEntry.worker_id == worker_id,
            models.LeaveEntry.leave_type.in_(PAID_LEAVE_TYPES),
            models.LeaveEntry.date_from <= end_date,
            models.LeaveEntry.date_to >= start_date,
        )
        .all()
    )
    dates: set[date] = set()
    for e in entries:
        d = max(e.date_from, start_date)
        last = min(e.date_to, end_date)
        while d <= last:
            dates.add(d)
            d += timedelta(days=1)
    return dates


def _monthly_attendance_rows(
    db: Session, worker_id: int, start_date: date, end_date: date, shifts: dict[str, models.ShiftConfig]
) -> list[dict]:
    """One dict per calendar day in the range -- the single source of
    truth both Form 25 (days as columns) and Form 25-B (days as rows)
    read from, so they can never silently disagree with each other."""
    records = (
        db.query(models.Attendance)
        .filter(
            models.Attendance.worker_id == worker_id,
            models.Attendance.date >= start_date,
            models.Attendance.date <= end_date,
        )
        .all()
    )
    by_date: dict[date, list[models.Attendance]] = {}
    for r in records:
        by_date.setdefault(r.date, []).append(r)
    paid_leave = _paid_leave_dates(db, worker_id, start_date, end_date)

    rows = []
    d = start_date
    while d <= end_date:
        day_records = by_date.get(d, [])
        present = [r for r in day_records if r.status == "present"]
        total_hours = sum(_shift_duration_hours(shifts.get(r.slot)) + (r.overtime_hours or 0) for r in present)
        day_paid_leave = d in paid_leave
        rows.append(
            {
                "date": d,
                "present": bool(present),
                "absent_marked": bool(day_records) and not present,
                "slots_present": sorted({r.slot for r in present}),
                "ot_hours": sum(r.overtime_hours or 0 for r in present),
                "total_hours": total_hours,
                "paid_leave": day_paid_leave,
                "leave_only": day_paid_leave and not present,
                "sunday": d.weekday() == 6,
            }
        )
        d += timedelta(days=1)
    return rows


def _summarize_month(rows: list[dict]) -> dict:
    days_worked = sum(1 for r in rows if r["present"])
    total_hours = sum(r["total_hours"] for r in rows)
    ot_hours = sum(r["ot_hours"] for r in rows)
    days_absent = sum(1 for r in rows if not r["present"])
    # Loss of pay: not present and not covered by paid leave. An
    # unmarked day (no attendance record at all) is treated the same as
    # an explicitly-marked absence -- the app has no separate "not yet
    # marked" state, so this is the only consistent reading available.
    loss_of_pay_days = sum(1 for r in rows if not r["present"] and not r["paid_leave"])
    # "Counted for wages incl. weekly holidays": present, paid leave, or
    # a Sunday (hardcoded as the weekly holiday for this phase -- see
    # Non-goals in PHASE3_STATUTORY_FORMS_PLAN.md).
    counted_for_wages = sum(1 for r in rows if r["present"] or r["paid_leave"] or r["sunday"])
    return {
        "days_worked": days_worked,
        "total_hours": total_hours,
        "ot_hours": ot_hours,
        "days_absent": days_absent,
        "loss_of_pay_days": loss_of_pay_days,
        "counted_for_wages": counted_for_wages,
    }


def _all_workers(db: Session, owner_id: int) -> list[models.Worker]:
    return db.query(models.Worker).filter(models.Worker.owner_id == owner_id).order_by(models.Worker.created_at).all()


def _form12_serial_map(db: Session, owner_id: int) -> dict[int, int]:
    """Every register numbers a worker by their row position in Form 12
    itself (registration order, oldest first, exit or not -- a real
    register never renumbers or drops a row) -- Form 25 and Form 15 both
    have a column that cross-references this exact number, so it must
    come from one shared place."""
    return {w.id: i for i, w in enumerate(_all_workers(db, owner_id), start=1)}


def _date_of_480_days_service(db: Session, worker_id: int) -> date | None:
    """The statutory trigger is 480 days actually *worked*, not 480
    calendar days since joining -- computed from real Attendance history
    (distinct present dates) rather than approximated, per this app's
    no-fabricated-figures rule. None until (if ever) that 480th present
    day has actually occurred."""
    present_dates = (
        db.query(models.Attendance.date)
        .filter(models.Attendance.worker_id == worker_id, models.Attendance.status == "present")
        .distinct()
        .order_by(models.Attendance.date)
        .all()
    )
    if len(present_dates) < 480:
        return None
    return present_dates[479][0]


def _compliance_map(db: Session, owner_id: int) -> dict[int, models.WorkerCompliance]:
    rows = (
        db.query(models.WorkerCompliance)
        .join(models.Worker, models.Worker.id == models.WorkerCompliance.worker_id)
        .filter(models.Worker.owner_id == owner_id)
        .all()
    )
    return {c.worker_id: c for c in rows}


def _person_counts(db: Session, owner_id: int) -> dict[str, int]:
    rows = (
        db.query(models.Worker.gender, models.WorkerCompliance.category)
        .outerjoin(models.WorkerCompliance, models.WorkerCompliance.worker_id == models.Worker.id)
        .filter(models.Worker.owner_id == owner_id)
        .all()
    )
    counts = {"men": 0, "women": 0, "male_adolescent": 0, "female_adolescent": 0}
    for gender, category in rows:
        is_young = category == "young_person"
        g = (gender or "").strip().lower()
        if g == "female":
            counts["female_adolescent" if is_young else "women"] += 1
        else:
            # Male, or a gender ("Other"/unset) that doesn't map onto
            # Form 15's 4 fixed statutory buckets -- counted here rather
            # than silently dropped from the total. A real limitation of
            # the government form itself, not something to engineer
            # around (see PHASE3_STATUTORY_FORMS_PLAN.md's review notes).
            counts["male_adolescent" if is_young else "men"] += 1
    return counts


def _wage_rate_as_of(db: Session, worker_id: int, as_of: date) -> models.WageProfile | None:
    return (
        db.query(models.WageProfile)
        .filter(models.WageProfile.worker_id == worker_id, models.WageProfile.effective_from <= as_of)
        .order_by(models.WageProfile.effective_from.desc())
        .first()
    )


def _leave_wages_for_period(db: Session, worker_id: int, start_date: date, end_date: date) -> float:
    entries = (
        db.query(models.LeaveEntry)
        .filter(
            models.LeaveEntry.worker_id == worker_id,
            models.LeaveEntry.leave_type.in_(PAID_LEAVE_TYPES),
            models.LeaveEntry.date_from <= end_date,
            models.LeaveEntry.date_to >= start_date,
        )
        .all()
    )
    return sum(e.wages_paid or 0 for e in entries)


def compute_wage(db: Session, owner_id: int, worker_id: int, month: int, year: int) -> dict | None:
    """The one place Form 15 and the Wage Slip's numbers come from --
    returns None if no WageProfile has ever been set for this worker
    (nothing to compute), never a fabricated zero."""
    start_date, end_date = _month_date_range(month, year)
    rate = _wage_rate_as_of(db, worker_id, end_date)
    if rate is None:
        return None

    shifts = _shift_map(db, owner_id)
    rows = _monthly_attendance_rows(db, worker_id, start_date, end_date, shifts)
    summary = _summarize_month(rows)

    if rate.rate_type == "daily":
        basic_wage = rate.basic * summary["days_worked"]
        hourly_rate = rate.basic / STANDARD_DAILY_HOURS
    else:
        days_in_month = (end_date - start_date).days + 1
        basic_wage = rate.basic
        hourly_rate = (rate.basic / days_in_month) / STANDARD_DAILY_HOURS

    ot_wages = summary["ot_hours"] * hourly_rate * OT_MULTIPLIER
    leave_wages = _leave_wages_for_period(db, worker_id, start_date, end_date)
    gross = basic_wage + rate.da + rate.hra + rate.other_allowances + ot_wages + leave_wages

    # PF wage base is Basic+DA (not Basic alone, not Gross); ESI wage
    # base is Gross. Standard statutory definitions, but the percentages
    # themselves are owner-entered on WageProfile, never hardcoded here.
    pf = rate.pf_rate / 100 * (basic_wage + rate.da)
    esi = rate.esi_rate / 100 * gross
    lwf = rate.lwf_amount
    # Advances and damages/fines ledgers are out of scope this phase
    # (see Non-goals) -- they contribute 0, not a fabricated figure.
    total_deductions = pf + esi + lwf
    net = gross - total_deductions

    payment = (
        db.query(models.WagePayment)
        .filter(models.WagePayment.worker_id == worker_id, models.WagePayment.month == month, models.WagePayment.year == year)
        .first()
    )

    return {
        "rate": rate,
        "summary": summary,
        "basic_wage": basic_wage,
        "ot_wages": ot_wages,
        "leave_wages": leave_wages,
        "gross": gross,
        "pf": pf,
        "esi": esi,
        "lwf": lwf,
        "total_deductions": total_deductions,
        "net": net,
        "payment": payment,
    }


def _header_elements(owner: models.Owner, styles, form_title: str, period_label: str | None = None) -> list:
    elements = [Paragraph(form_title, styles["Title"])]
    elements.append(Paragraph(owner.factory_name, styles["Normal"]))
    if owner.factory_address:
        elements.append(Paragraph(owner.factory_address, styles["Normal"]))
    if owner.factory_licence_no:
        elements.append(Paragraph(f"Licence / Registration No.: {owner.factory_licence_no}", styles["Normal"]))
    if period_label:
        elements.append(Paragraph(period_label, styles["Normal"]))
    elements.append(Spacer(1, 0.4 * cm))
    return elements


_SMALL_CELL_STYLE = ParagraphStyle("small_cell", fontName="Helvetica", fontSize=6.5, leading=8)
_SMALL_HEADER_STYLE = ParagraphStyle("small_header", fontName="Helvetica-Bold", fontSize=6.5, leading=8, textColor=pdf_colors.white)


def _wrap_row(values: list, header: bool = False) -> list:
    """Wraps every cell in a Paragraph so long header labels and
    free-text values (addresses, bank details) wrap within their column
    instead of overflowing -- plain strings in a reportlab Table never
    wrap on their own."""
    style = _SMALL_HEADER_STYLE if header else _SMALL_CELL_STYLE
    return [Paragraph(str(v), style) for v in values]


def _style_table(table: Table) -> None:
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), pdf_colors.HexColor("#1B2340")),
                ("TEXTCOLOR", (0, 0), (-1, 0), pdf_colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                ("GRID", (0, 0), (-1, -1), 0.4, pdf_colors.HexColor("#E5E7EB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [pdf_colors.white, pdf_colors.HexColor("#F4F6F9")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )


# Landscape A3 -- an actual standard, printable paper size (unlike the
# 50-85cm custom pages these registers used before, which no printer can
# handle: the print driver either clips columns past its real paper
# width or shrinks the whole page down until it's illegible). 38cm is
# the safe usable width inside 1cm margins with a little headroom for
# cell padding/borders.
REGISTER_PAGESIZE = landscape(A3)
REGISTER_PAGE_BUDGET_CM = 38.0


def _paginated_register_elements(
    header: list[str],
    rows: list[list[str]],
    col_widths_cm: list[float],
    fixed_count: int,
    page_budget_cm: float = REGISTER_PAGE_BUDGET_CM,
) -> list:
    """Splits a register with more columns than fit on one printable
    page into column groups -- the same way the real paper registers
    are split across a fold or facing pages, not something invented
    here. The first `fixed_count` columns (serial number, name, ...)
    repeat in every group so a printed page is self-identifying on its
    own, and each group after the first starts on a new page."""
    fixed_widths_cm = col_widths_cm[:fixed_count]
    fixed_sum_cm = sum(fixed_widths_cm)
    groups: list[list[int]] = []
    current: list[int] = []
    current_sum_cm = fixed_sum_cm
    for i in range(fixed_count, len(col_widths_cm)):
        w = col_widths_cm[i]
        if current and current_sum_cm + w > page_budget_cm:
            groups.append(current)
            current = []
            current_sum_cm = fixed_sum_cm
        current.append(i)
        current_sum_cm += w
    if current:
        groups.append(current)
    if not groups:
        groups = [[]]

    elements: list = []
    for group_index, group in enumerate(groups):
        indices = list(range(fixed_count)) + group
        group_header = [header[i] for i in indices]
        group_widths = [col_widths_cm[i] * cm for i in indices]
        group_rows = [[row[i] for i in indices] for row in rows]
        table_data = [group_header] + group_rows
        wrapped_data = [_wrap_row(table_data[0], header=True)] + [_wrap_row(r) for r in table_data[1:]]
        table = Table(wrapped_data, colWidths=group_widths, repeatRows=1)
        _style_table(table)
        if group_index > 0:
            elements.append(PageBreak())
            elements.append(
                Paragraph(f"(continued -- column group {group_index + 1} of {len(groups)})", getSampleStyleSheet()["Normal"])
            )
            elements.append(Spacer(1, 0.2 * cm))
        elements.append(table)
    return elements


# --------------------------------------------------------------------------
# Form 25 -- Muster Roll and Register (factory-wide, monthly)
# --------------------------------------------------------------------------


FORM25_HEADER = [
    "Sl.No.",
    "Sl.No. in Register of Adult Workers and Young Persons",
    "Name of the Worker",
    "Workers Identity Number",
    "Time at which work commenced",
    "Rest Interval",
    "Time at which work ends",
    "Scheme of Shifts",
]
FORM25_TRAILER = [
    "Total Days Worked",
    "Total Hours Worked",
    "No. of Days on Loss of Pay",
    "Benefits Availed for Working on National Holiday",
    "Benefits Availed for Working on Festival Holiday",
    "Remarks",
]


def _form25_table_data(db: Session, owner: models.Owner, month: int, year: int) -> tuple[list, date, date]:
    """Column order matches Form 25 -- Muster Roll and Register exactly,
    fields (1) through (15) per the real Tamil Nadu form: Sl.No,
    cross-reference to Form 12's own Sl.No, Name, Worker ID, the shift's
    commenced/rest/end times and scheme, one column per calendar day
    (daily hours of work including overtime), then Total Days Worked,
    Total Hours Worked, Loss of Pay days, the two Holiday-benefit
    columns, and Remarks. Benefits-on-holiday aren't tracked anywhere in
    this app yet (no holiday register exists) so those two columns are
    always "-" rather than a fabricated figure."""
    start_date, end_date = _month_date_range(month, year)
    workers = _all_workers(db, owner.id)
    shifts = _shift_map(db, owner.id)
    compliance = _compliance_map(db, owner.id)
    serials = _form12_serial_map(db, owner.id)
    days_in_month = end_date.day

    header = FORM25_HEADER + [str(d) for d in range(1, days_in_month + 1)] + FORM25_TRAILER
    table_data = [header]
    for i, worker in enumerate(workers, start=1):
        rows = _monthly_attendance_rows(db, worker.id, start_date, end_date, shifts)
        summary = _summarize_month(rows)
        used_slots = sorted({s for r in rows for s in r["slots_present"]})
        used_shifts = [shifts[k] for k in used_slots if k in shifts]
        shift_labels = ", ".join(s.label for s in used_shifts) or "-"
        commenced = ", ".join(s.start_time for s in used_shifts if s.start_time) or "-"
        ends = ", ".join(s.end_time for s in used_shifts if s.end_time) or "-"
        rest = ", ".join(s.rest_interval for s in used_shifts if s.rest_interval) or "-"
        day_cells = [
            f"{r['total_hours']:.1f}" if r["present"] else ("L" if r["leave_only"] else "-") for r in rows
        ]
        worker_code = compliance.get(worker.id).worker_code if compliance.get(worker.id) else None
        table_data.append(
            [
                str(i),
                str(serials.get(worker.id, "-")),
                worker.name,
                worker_code or "-",
                commenced,
                rest,
                ends,
                shift_labels,
            ]
            + day_cells
            + [
                str(summary["days_worked"]),
                f"{summary['total_hours']:.1f}",
                str(summary["loss_of_pay_days"]),
                "-",
                "-",
                "-",
            ]
        )
    return table_data, start_date, end_date


def build_form25(db: Session, owner: models.Owner, month: int, year: int, format: str) -> tuple[bytes, str, str]:
    table_data, start_date, end_date = _form25_table_data(db, owner, month, year)
    period_label = f"{start_date.isoformat()} to {end_date.isoformat()}"

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Form 25"
        ws.append([owner.factory_name, period_label])
        for row in table_data:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return (
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"form25_{year}_{month:02d}.xlsx",
        )

    # Column widths -- without them reportlab auto-splits the page width
    # evenly across all ~45 columns, so even "Name of the Worker" ends up
    # a couple of centimetres wide and every word wraps onto its own
    # line. Day columns only ever hold a couple of characters ("8.0",
    # "-", "L"); the named/id/time columns need more. There are too many
    # columns to fit on one printable page, so they're split into groups
    # below (_paginated_register_elements) rather than one custom
    # oversized page a real printer can't handle.
    header_widths_cm = [1.0, 2.2, 2.8, 1.8, 1.8, 1.6, 1.8, 1.8]
    day_width_cm = 0.85
    trailer_widths_cm = [1.7, 1.7, 1.6, 2.1, 2.1, 1.7]
    days_in_month = end_date.day
    col_widths_cm = header_widths_cm + [day_width_cm] * days_in_month + trailer_widths_cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=REGISTER_PAGESIZE, topMargin=1 * cm, bottomMargin=1 * cm, leftMargin=1 * cm, rightMargin=1 * cm
    )
    styles = getSampleStyleSheet()
    elements = _header_elements(owner, styles, "Form 25 -- Muster Roll and Register", period_label)
    elements.append(
        Paragraph(
            "Register of Compensatory Holidays, and Benefits Availed for Working on National/Festival Holiday: "
            "not tracked this phase -- no Holiday register exists yet (see PHASE3_STATUTORY_FORMS_PLAN.md's Non-goals).",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 0.3 * cm))
    # Sl.No, Sl.No-in-Register, Name, Worker ID repeat on every printed
    # page so each one still identifies whose row it is on its own.
    elements.extend(_paginated_register_elements(table_data[0], table_data[1:], col_widths_cm, fixed_count=4))
    doc.build(elements)
    return buf.getvalue(), "application/pdf", f"form25_{year}_{month:02d}.pdf"


# --------------------------------------------------------------------------
# Form 25-B -- Time Card (per worker, monthly)
# --------------------------------------------------------------------------


def build_form25b(db: Session, owner: models.Owner, worker: models.Worker, month: int, year: int, format: str) -> tuple[bytes, str, str]:
    """Field order follows Form 25-B's own field list: the header block
    (factory, licence, worker, father's name, ticket/token no.,
    designation, date of entry), a day-wise table (date, in/out time,
    rest interval, OT hours, total hours), then the four monthly
    summary counts and the manager's signature line. No scanned copy of
    this form was provided (only Form 12/25/15/Wage Slip were), so this
    matches the field list rather than a pixel-exact layout."""
    start_date, end_date = _month_date_range(month, year)
    shifts = _shift_map(db, owner.id)
    compliance = db.query(models.WorkerCompliance).filter(models.WorkerCompliance.worker_id == worker.id).first()
    rows = _monthly_attendance_rows(db, worker.id, start_date, end_date, shifts)
    summary = _summarize_month(rows)
    period_label = f"{start_date.isoformat()} to {end_date.isoformat()}"

    table_data = [["Date", "In Time", "Out Time", "Interval", "OT Hours Worked", "Total Hours Worked"]]
    for r in rows:
        used = [shifts[s] for s in r["slots_present"] if s in shifts]
        table_data.append(
            [
                r["date"].isoformat(),
                ", ".join(s.start_time for s in used if s.start_time) or "-",
                ", ".join(s.end_time for s in used if s.end_time) or "-",
                ", ".join(s.rest_interval for s in used if s.rest_interval) or "-",
                f"{r['ot_hours']:.1f}",
                f"{r['total_hours']:.1f}",
            ]
        )

    summary_rows = [
        ("No. of Days Attendance During the Month", str(summary["days_worked"])),
        ("No. of Days Absent", str(summary["days_absent"])),
        ("No. of Days of Leave Granted with Wages", str(sum(1 for r in rows if r["paid_leave"]))),
        ("No. of Days Counted for Wages Including Weekly Holidays", str(summary["counted_for_wages"])),
    ]

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Form 25-B"
        ws.append([owner.factory_name, owner.factory_licence_no or "-", period_label])
        ws.append(
            [
                worker.name,
                compliance.father_or_spouse_name if compliance else "-",
                compliance.worker_code if compliance else "-",
                compliance.designation_or_nature_of_work if compliance else "-",
                compliance.date_of_joining.isoformat() if compliance and compliance.date_of_joining else "-",
            ]
        )
        for row in table_data:
            ws.append(row)
        ws.append([])
        for label, value in summary_rows:
            ws.append([label, value])
        buf = io.BytesIO()
        wb.save(buf)
        return (
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"form25b_{worker.id}_{year}_{month:02d}.xlsx",
        )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = _header_elements(owner, styles, "Form 25-B -- Time Card", period_label)
    elements.append(Paragraph(f"Name of the Worker: {worker.name}", styles["Normal"]))
    elements.append(Paragraph(f"Father's Name: {compliance.father_or_spouse_name if compliance else '-'}", styles["Normal"]))
    elements.append(Paragraph(f"Ticket No. / Token No.: {compliance.worker_code if compliance else '-'}", styles["Normal"]))
    elements.append(
        Paragraph(f"Designation or Occupation: {compliance.designation_or_nature_of_work if compliance else '-'}", styles["Normal"])
    )
    elements.append(
        Paragraph(
            f"Date of Entry into Service: {compliance.date_of_joining.isoformat() if compliance and compliance.date_of_joining else '-'}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 0.3 * cm))
    table = Table(table_data, repeatRows=1)
    _style_table(table)
    elements.append(table)
    elements.append(Spacer(1, 0.4 * cm))
    for label, value in summary_rows:
        elements.append(Paragraph(f"{label}: {value}", styles["Normal"]))
    elements.append(Spacer(1, 0.6 * cm))
    elements.append(Paragraph("Date & Signature of the Manager: ____________________", styles["Normal"]))
    doc.build(elements)
    return buf.getvalue(), "application/pdf", f"form25b_{worker.id}_{year}_{month:02d}.pdf"


# --------------------------------------------------------------------------
# Form 12 -- Register of Adult Workers & Young Persons (one-time, per worker)
# --------------------------------------------------------------------------


FORM12_HEADER = [
    "Serial Number",
    "Name of the Worker",
    "Working Identity No.",
    "Gender",
    "Father / Spouse Name",
    "Date of Birth",
    "Present Address",
    "Permanent Address",
    "Aadhaar No.",
    "Date of entry into Service",
    "Designation / Nature of work",
    "EPF No. / UAN No.",
    "ESIC No.",
    "Date on which completion of 480 days of Service",
    "Date on which made permanent",
    "Period of Suspension, if any",
    "Bank A/c No., Name of Bank, Branch (IFSC)",
    "Photo",
    "Mobile Number",
    "E-mail ID",
    "Specimen Signature / Thumb Impression",
    "Date of exit",
    "Reason to exit",
    "Remarks",
]


def _form12_row(worker: models.Worker, compliance: models.WorkerCompliance | None, sl_no: int, db: Session) -> list[str]:
    """One register row, all 24 columns in the real Form 12's own order.
    Photo, Specimen Signature, and E-mail ID have no capture mechanism
    anywhere in this app (no file storage, no email field on Worker) --
    left blank rather than a fabricated placeholder. Present/Permanent
    address map onto this app's current/native address, the closest
    existing fields; Aadhaar stays masked to last-4 (existing privacy
    practice), never decrypted for a printed register."""
    bank_parts = [p for p in [worker.bank_account_number, worker.bank_ifsc] if p]
    date_480 = _date_of_480_days_service(db, worker.id)
    return [
        str(sl_no),
        worker.name,
        compliance.worker_code if compliance and compliance.worker_code else "-",
        worker.gender or "-",
        compliance.father_or_spouse_name if compliance and compliance.father_or_spouse_name else "-",
        worker.dob.isoformat() if worker.dob else "-",
        worker.current_address or "-",
        worker.native_address or "-",
        f"XXXX XXXX {worker.aadhaar_last4}",
        compliance.date_of_joining.isoformat() if compliance and compliance.date_of_joining else "-",
        compliance.designation_or_nature_of_work if compliance and compliance.designation_or_nature_of_work else "-",
        compliance.epf_uan_no if compliance and compliance.epf_uan_no else "-",
        compliance.esic_no if compliance and compliance.esic_no else "-",
        date_480.isoformat() if date_480 else "-",
        compliance.date_made_permanent.isoformat() if compliance and compliance.date_made_permanent else "-",
        compliance.suspension_period if compliance and compliance.suspension_period else "-",
        ", ".join(bank_parts) if bank_parts else "-",
        "-",
        worker.mobile or "-",
        "-",
        "-",
        worker.deactivated_at.date().isoformat() if worker.deactivated_at else "-",
        worker.deactivated_reason or "-",
        "-",
    ]


def build_form12(db: Session, owner: models.Owner, format: str, worker: models.Worker | None = None) -> tuple[bytes, str, str]:
    """Form 12 -- Register of Adult Workers and Young Persons. The real
    government form is a running register (every worker who has ever
    been employed is a row, in registration order, never removed on
    exit -- that's what the Date of exit/Reason to exit columns are
    for), not a per-worker sheet -- confirmed against the actual scanned
    form. `worker` narrows the register to a single row (used by the
    per-worker download); omit it for the full factory register."""
    compliance = _compliance_map(db, owner.id)
    if worker is not None:
        workers = [worker]
        serials = _form12_serial_map(db, owner.id)
        rows = [_form12_row(worker, compliance.get(worker.id), serials.get(worker.id, 1), db)]
        filename_suffix = f"_{worker.id}"
    else:
        workers = _all_workers(db, owner.id)
        rows = [_form12_row(w, compliance.get(w.id), i, db) for i, w in enumerate(workers, start=1)]
        filename_suffix = ""
    table_data = [FORM12_HEADER] + rows

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Form 12"
        ws.append([owner.factory_name, f"Registration No.: {owner.factory_licence_no or '-'}"])
        for row in table_data:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"form12{filename_suffix}.xlsx"

    # Explicit widths (cm), one per column in FORM12_HEADER's order --
    # same reasoning as Form 25: unsized columns split the page evenly
    # and even short values like names or dates end up wrapping.
    col_widths_cm = [
        1.0, 2.8, 1.8, 1.2, 2.8, 1.6, 4.0, 4.0, 2.2, 1.8,
        2.4, 2.0, 1.6, 1.8, 1.8, 2.0, 3.6, 1.4, 1.8, 2.2,
        2.4, 1.6, 2.2, 2.0,
    ]

    buf = io.BytesIO()
    # 24 columns, several of them free-text addresses/bank details -- too
    # many to fit one printable page, split into groups below.
    doc = SimpleDocTemplate(
        buf, pagesize=REGISTER_PAGESIZE, topMargin=1 * cm, bottomMargin=1 * cm, leftMargin=1 * cm, rightMargin=1 * cm
    )
    styles = getSampleStyleSheet()
    elements = _header_elements(owner, styles, "Form 12 -- Register of Adult Workers and Young Persons")
    elements.append(Paragraph(f"Registration No.: {owner.factory_licence_no or '-'}", styles["Normal"]))
    elements.append(Spacer(1, 0.3 * cm))
    # Serial Number + Name of the Worker repeat on every printed page.
    elements.extend(_paginated_register_elements(table_data[0], table_data[1:], col_widths_cm, fixed_count=2))
    doc.build(elements)
    return buf.getvalue(), "application/pdf", f"form12{filename_suffix}.pdf"


# --------------------------------------------------------------------------
# Form 15 -- Register of Leave with Wages, Part II (factory-wide, monthly)
# --------------------------------------------------------------------------


FORM15_HEADER = [
    "Serial Number",
    "Sl.No. in Register of Adult Workers and Young Persons",
    "Name of the Worker",
    "Worked Identity Number",
    "Number of Days Worked",
    "Basic Wages",
    "Dearness Allowance",
    "House Rent Allowance",
    "Other Allowance (nature may be specified)",
    "Overtime Wages",
    "Leave Wages (Earned leave / National, Festival & Special Holidays / Others)",
    "Gross Wages",
    "Provident Fund",
    "Employees State Insurance",
    "Labour Welfare Fund",
    "Advance Paid",
    "Advance Recovery pending at the beginning of the month",
    "Advance Recovery",
    "Pending Recovery",
    "Deduction imposed on Damages, Loss or Fines",
    "Deduction recovery pending at the beginning of the month",
    "Deduction made on Damages, Loss or Fines",
    "Pending Recovery",
    "Net Wages",
    "Date of Payment",
    "Unpaid accumulation",
    "Rate at which subsistence allowance calculated and amount paid",
    "Receipt by worker",
    "Bank Transaction ID and Date",
    "Remarks",
]


def build_form15(db: Session, owner: models.Owner, month: int, year: int, format: str) -> tuple[bytes, str, str]:
    """Full 30-column layout per the real Form 15 -- Register of Leave
    with Wages (Part II). Advances and damages/fines ledgers aren't
    tracked anywhere in this app (see Non-goals) so those columns are
    always "0.00", never a fabricated figure; likewise Unpaid
    Accumulation, Subsistence Allowance, and Receipt by Worker have no
    backing data and stay "-"."""
    start_date, end_date = _month_date_range(month, year)
    workers = _all_workers(db, owner.id)
    compliance = _compliance_map(db, owner.id)
    counts = _person_counts(db, owner.id)
    serials = _form12_serial_map(db, owner.id)
    period_label = f"{start_date.isoformat()} to {end_date.isoformat()}"

    table_data = [FORM15_HEADER]
    for i, worker in enumerate(workers, start=1):
        wage = compute_wage(db, owner.id, worker.id, month, year)
        worker_code = compliance.get(worker.id).worker_code if compliance.get(worker.id) else None
        sl_no = str(serials.get(worker.id, "-"))
        if wage is None:
            table_data.append([str(i), sl_no, worker.name, worker_code or "-", "-"] + ["no wage rate set"] * 25)
            continue
        payment_ref = wage["payment"].payment_reference if wage["payment"] else None
        payment_date = wage["payment"].date_of_payment if wage["payment"] else None
        bank_txn = f"{payment_ref} / {payment_date.isoformat()}" if payment_ref and payment_date else "-"
        table_data.append(
            [
                str(i),
                sl_no,
                worker.name,
                worker_code or "-",
                str(wage["summary"]["days_worked"]),
                f"{wage['basic_wage']:.2f}",
                f"{wage['rate'].da:.2f}",
                f"{wage['rate'].hra:.2f}",
                f"{wage['rate'].other_allowances:.2f}",
                f"{wage['ot_wages']:.2f}",
                f"{wage['leave_wages']:.2f}",
                f"{wage['gross']:.2f}",
                f"{wage['pf']:.2f}",
                f"{wage['esi']:.2f}",
                f"{wage['lwf']:.2f}",
                "0.00",  # Advance Paid -- no advances ledger this phase (Non-goals)
                "0.00",  # Advance Recovery pending at beginning of month
                "0.00",  # Advance Recovery
                "0.00",  # Pending Recovery (advances)
                "0.00",  # Deduction imposed on Damages, Loss or Fines -- no ledger this phase
                "0.00",  # Deduction recovery pending at beginning of month
                "0.00",  # Deduction made on Damages, Loss or Fines
                "0.00",  # Pending Recovery (damages)
                f"{wage['net']:.2f}",
                payment_date.isoformat() if payment_date else "-",
                "-",  # Unpaid accumulation -- no ledger this phase
                "-",  # Rate at which subsistence allowance calculated and amount paid
                "-",  # Receipt by worker -- no signature/receipt capture this phase
                bank_txn,
                "-",
            ]
        )

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Form 15"
        ws.append([owner.factory_name, period_label])
        ws.append(
            ["Men", counts["men"], "Women", counts["women"], "Male Adolescent", counts["male_adolescent"], "Female Adolescent", counts["female_adolescent"]]
        )
        for row in table_data:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"form15_{year}_{month:02d}.xlsx"

    # Explicit widths (cm), one per column in FORM15_HEADER's order --
    # same reasoning as Form 25/12.
    col_widths_cm = [
        0.9, 1.8, 2.6, 1.8, 1.6, 1.6, 1.6, 1.6, 1.8, 1.6,
        2.4, 1.6, 1.6, 1.8, 1.6, 1.6, 2.2, 1.6, 1.6, 2.2,
        2.2, 2.2, 1.6, 1.6, 1.6, 1.8, 2.4, 1.6, 2.2, 1.6,
    ]

    buf = io.BytesIO()
    # Same standard printable page as Form 25/12, split into column
    # groups below for the same reason.
    doc = SimpleDocTemplate(
        buf, pagesize=REGISTER_PAGESIZE, topMargin=1 * cm, bottomMargin=1 * cm, leftMargin=1 * cm, rightMargin=1 * cm
    )
    styles = getSampleStyleSheet()
    elements = _header_elements(owner, styles, "Form 15 -- Register of Leave with Wages (Part II)", period_label)
    counts_table = Table(
        [["Men", "Women", "Male Adolescent", "Female Adolescent"], [counts["men"], counts["women"], counts["male_adolescent"], counts["female_adolescent"]]],
        colWidths=[3.5 * cm] * 4,
    )
    counts_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), pdf_colors.HexColor("#1B2340")),
                ("TEXTCOLOR", (0, 0), (-1, 0), pdf_colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, pdf_colors.HexColor("#E5E7EB")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]
        )
    )
    elements.append(Paragraph("Total Number of Persons Employed:", styles["Normal"]))
    elements.append(counts_table)
    elements.append(Spacer(1, 0.3 * cm))
    # Serial Number, Sl.No-in-Register, Name, Worker ID repeat on every
    # printed page.
    elements.extend(_paginated_register_elements(table_data[0], table_data[1:], col_widths_cm, fixed_count=4))
    doc.build(elements)
    return buf.getvalue(), "application/pdf", f"form15_{year}_{month:02d}.pdf"


# --------------------------------------------------------------------------
# Wage Slip -- per worker, monthly (formatted subset of Form 15's numbers)
# --------------------------------------------------------------------------


def build_wageslip(db: Session, owner: models.Owner, worker: models.Worker, month: int, year: int, format: str) -> tuple[bytes, str, str]:
    """Field order/labels follow the "Fields of Wage Slip" list exactly
    (the Tamil chit-style sample sent alongside it is a different,
    unrelated form and is intentionally not used as a reference here,
    per instruction). Deductions are broken out as the field list's own
    a./b. sub-items under a "Deductions" heading row; Labour Welfare
    Fund has no dedicated line on this form so it's folded into "Other
    Deductions", the field list's catch-all."""
    wage = compute_wage(db, owner.id, worker.id, month, year)
    compliance = db.query(models.WorkerCompliance).filter(models.WorkerCompliance.worker_id == worker.id).first()
    period_label = f"{year}-{month:02d}"

    if wage is None:
        rows = [["Status", "No wage rate has been set for this worker yet"]]
    else:
        rows = [
            ["Wage Slip No. / Worker ID No.", compliance.worker_code if compliance else "-"],
            ["Name of the Worker", worker.name],
            ["Nature of Work / Designation", compliance.designation_or_nature_of_work if compliance else "-"],
            ["Wage Period", period_label],
            ["Minimum Wages / day or month", f"{wage['rate'].basic:.2f} / {wage['rate'].rate_type}"],
            ["Total Days Worked", str(wage["summary"]["days_worked"])],
            ["Gross Wages", f"{wage['gross']:.2f}"],
            ["Deductions", ""],
            ["   a. Provident Fund", f"{wage['pf']:.2f}"],
            ["   b. Employees State Insurance", f"{wage['esi']:.2f}"],
            ["Other Deductions", f"{wage['lwf']:.2f}"],
            ["Net Wages Paid", f"{wage['net']:.2f}"],
            [
                "Date of Payment",
                wage["payment"].date_of_payment.isoformat() if wage["payment"] and wage["payment"].date_of_payment else "-",
            ],
        ]

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Wage Slip"
        ws.append([owner.factory_name])
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"wageslip_{worker.id}_{year}_{month:02d}.xlsx"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = _header_elements(owner, styles, "Wage Slip", period_label)
    table = Table(rows, colWidths=[7 * cm, 9 * cm])
    _style_table(table)
    elements.append(table)
    elements.append(Spacer(1, 0.6 * cm))
    elements.append(Paragraph("Signature / Thumb Impression of the Worker: ____________________", styles["Normal"]))
    elements.append(Paragraph("Manager's Signature: ____________________", styles["Normal"]))
    doc.build(elements)
    return buf.getvalue(), "application/pdf", f"wageslip_{worker.id}_{year}_{month:02d}.pdf"
